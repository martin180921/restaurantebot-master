"""Vista de Menú: administración de las secciones de la carta + ajustes.

Pestañas (la misma taxonomía que ven el POS y la app del cliente):
  🍽️ Plato del Día  → componentes por grupo (entrada / principio / proteína /
                       acompañamientos). Cada opción se activa, se marca "86" (agotado
                       hoy), se renombra o se elimina. Precio plano editable en Ajustes.
  (categorías)      → una pestaña por cada fila ACTIVA de la tabla 'categorias' (tabla
                       de datos, ver db.cargar_categorias), en su 'orden'. Las 4
                       clásicas (⭐ Especiales, 📋 A la carta, 🍟 Adicionales, 🥤 Bebidas)
                       vienen sembradas; el restaurante agrega las suyas (Desayunos,
                       Postres…) desde ⚙️ Ajustes → 🏷️ Categorías. Cada una es un
                       catálogo con su propio precio; solo Especiales/Adicionales
                       llevan descripción (ver db.comportamiento_categoria).
  ⚙️ Ajustes        → precio plano del Plato del Día, recargo de entrega
                       (Domicilio / Para Llevar), nº de acompañamientos y el editor de
                       categorías.

Reusa los patrones del panel: tarjetas (.menu-card), modales @st.dialog, toasts
db.flash() y la invalidación de caché (cargar_*.clear()) tras cada escritura.
"""
import streamlit as st
from sqlalchemy import text
from datetime import time as dtime
import html
import io
import csv
import json
import numbers
import re
import unicodedata
import pandas as pd

import auth
from db import (engine, titulo_seccion, cargar_menu, cargar_componentes, cargar_catalogo,
                cargar_ajustes, fmt_money, flash, disponibles,
                componentes_activos_por_grupo, precio_plato_dia,
                cargar_grupos_pd, etiquetas_grupos_pd, crear_grupo_pd,
                guardar_grupo_pd, eliminar_grupo_pd,
                cargar_categorias, etiquetas_categorias, comportamiento_categoria,
                extras_de_categoria, crear_categoria, guardar_categoria, eliminar_categoria,
                en_horario_categoria, ahora_bogota,
                guardar_inventario, stock_int, STOCK_BAJO, agotado_por_stock, hoy_bogota)


# Techo de INTEGER en Postgres. Un precio/stock por encima (celda gigante del Excel o
# fat-finger en un number_input) revienta el UPDATE con NumericValueOutOfRange: abortaba
# TODA la importación con un stacktrace, y en los editores del panel tumbaba la pantalla.
# Acotamos al rango válido en los dos caminos (ver _parse_int y los max_value de abajo).
_INT32_MAX = 2_147_483_647


# ── Helpers ──────────────────────────────────────────────────────────────────────
def _txt(valor) -> str:
    """Coerción segura a str para celdas que pueden venir None/NaN (descripcion)."""
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except (TypeError, ValueError):
        pass
    return str(valor)


def _agotado_hoy(valor) -> bool:
    """True si agotado_hasta cubre el día de hoy (patrón "86")."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return False
    try:
        return pd.Timestamp(valor).date() >= hoy_bogota()
    except (ValueError, TypeError):
        return False


def _clear_menu_caches():
    """Invalida las cachés que dependen de la tabla 'menu' (catálogo nuevo + lectura
    plana heredada que aún usa el POS hasta su rediseño)."""
    cargar_catalogo.clear()
    cargar_menu.clear()


def _stock_chip(stock) -> str:
    """Chip de existencias para las tarjetas: '📦 N' verde (>umbral) / ámbar (≤umbral) /
    rojo (0). Vacío si la opción no lleva control de stock (ilimitada). El stock diario
    se fija en la pestaña 📦 Inventario."""
    s = stock_int(stock)
    if s is None:
        return ""
    if s <= 0:
        bg, fg = "#fee2e2", "#991b1b"
    elif s <= STOCK_BAJO:
        bg, fg = "#fef3c7", "#92400e"
    else:
        bg, fg = "#dcfce7", "#15803d"
    return (f'<span class="badge" style="background:{bg}; color:{fg}; border:1px solid {bg};">'
            f'📦 {s}</span>')


# ── Disponibilidad y acordeón del menú (UI) ──────────────────────────────────────
def _disponible_hoy(row) -> bool:
    """True si la fila (plato o componente) se puede SERVIR hoy: activa, sin 86 manual
    del día y con stock (o ilimitada). Es el criterio para subirla a la sección superior
    'Disponibles' de su categoría; lo demás (inactivo / 86 / agotado) baja al fondo."""
    return (bool(row["activo"])
            and not _agotado_hoy(row.get("agotado_hasta"))
            and not agotado_por_stock(row.get("stock")))


def _partir_disponibles(sub):
    """(disponibles, no_disponibles) conservando el orden por 'orden' dentro de cada
    grupo (sub ya viene ordenado por cargar_catalogo()/cargar_componentes())."""
    disp, no = [], []
    if sub is not None and not sub.empty:
        for _, row in sub.iterrows():
            (disp if _disponible_hoy(row) else no).append(row)
    return disp, no


def _inject_accordion_css():
    """Estilo de las cabeceras de acordeón del Menú (botones con key 'acc_*'): ancho
    completo, alineadas a la izquierda y con aspecto de cabecera plegable (no de botón).
    Se inyecta DESPUÉS del CSS global de panel.py; la variante anclada en
    div[data-testid="stColumn"] iguala/gana la especificidad de las reglas de columna
    (mismo patrón que los botones semánticos del panel)."""
    st.markdown("""
    <style>
    [class*="st-key-acc_"] button,
    div[data-testid="stColumn"] [class*="st-key-acc_"] .stButton > button {
        text-align: left !important; justify-content: flex-start !important;
        background: #fbfbf9 !important; border: 1px solid #ececec !important;
        border-radius: 12px !important; padding: 12px 16px !important;
        font-size: 0.95rem !important; font-weight: 600 !important;
        color: #26262b !important; margin: 6px 0 !important; min-height: 0 !important;
    }
    [class*="st-key-acc_"] button:hover,
    div[data-testid="stColumn"] [class*="st-key-acc_"] .stButton > button:hover {
        background: #f2f1ed !important; border-color: #d8d6cf !important; color: #26262b !important;
    }
    [class*="st-key-acc_"] button p,
    div[data-testid="stColumn"] [class*="st-key-acc_"] .stButton > button p {
        text-align: left !important; margin: 0 !important; width: 100% !important; font-weight: 600 !important;
    }
    </style>
    """, unsafe_allow_html=True)


def _acc_header(key: str, label: str, resumen: str = "", *, default_open: bool = False) -> bool:
    """Cabecera de un acordeón de PLATOS dentro de una categoría (Disponibles /
    No disponibles). El estado abierto/cerrado vive en session_state['acc_open_<key>']
    para SOBREVIVIR a los st.rerun() de cada acción de la tarjeta (activar / 86 / editar /
    eliminar): st.expander reinicia su estado en cada rerun, por eso se gestiona a mano.
    default_open fija el estado inicial. Devuelve True si la sección queda desplegada."""
    open_key = f"acc_open_{key}"
    abierto = bool(st.session_state.get(open_key, default_open))
    chevron = "▾" if abierto else "▸"
    extra = f"   ·   {resumen}" if resumen else ""
    if st.button(f"{chevron}   {label}{extra}", key=f"acc_{key}", use_container_width=True):
        st.session_state[open_key] = not abierto
        st.rerun()
    return bool(st.session_state.get(open_key, default_open))


# ══════════════════════════════════════════════════════════════════════════════
# DB: componentes del Plato del Día (tabla menu_componentes)
# ══════════════════════════════════════════════════════════════════════════════
def agregar_componente(grupo: str, nombre: str):
    """Inserta una opción nueva en el grupo, salvo que ya exista una con el mismo
    nombre (sin distinguir mayúsculas/espacios) en ese grupo: dos componentes
    iguales chocan en el selector del POS (misma key de botón) y revientan la
    pantalla con StreamlitDuplicateElementKey."""
    nombre = nombre.strip()
    with engine.begin() as conn:
        ya_existe = conn.execute(text(
            "SELECT 1 FROM menu_componentes WHERE grupo = :g AND lower(nombre) = lower(:n)"
        ), {"g": grupo, "n": nombre}).first()
        if ya_existe:
            flash(f"'{nombre}' ya existe en este grupo", "⚠️")
            return
        mx = conn.execute(text(
            "SELECT COALESCE(MAX(orden),0) FROM menu_componentes WHERE grupo = :g"
        ), {"g": grupo}).scalar()
        conn.execute(text(
            "INSERT INTO menu_componentes (grupo, nombre, activo, orden) "
            "VALUES (:g, :n, TRUE, :o)"
        ), {"g": grupo, "n": nombre, "o": mx + 1})
    cargar_componentes.clear()
    flash("Opción agregada", "✅")


def actualizar_componente(cid: int, nombre: str):
    with engine.begin() as conn:
        conn.execute(text("UPDATE menu_componentes SET nombre = :n WHERE id = :id"),
                     {"n": nombre.strip(), "id": cid})
    cargar_componentes.clear()
    flash("Opción actualizada", "✅")


def toggle_componente(cid: int, activo_actual: bool):
    with engine.begin() as conn:
        conn.execute(text("UPDATE menu_componentes SET activo = :a WHERE id = :id"),
                     {"a": not activo_actual, "id": cid})
    cargar_componentes.clear()
    flash("Opción activada" if not activo_actual else "Opción desactivada",
          "▶️" if not activo_actual else "⏸️")


def eliminar_componente(cid: int):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM menu_componentes WHERE id = :id"), {"id": cid})
    cargar_componentes.clear()
    flash("Opción eliminada", "🗑️")


def marcar_agotado_componente(cid: int):
    with engine.begin() as conn:
        conn.execute(text("UPDATE menu_componentes SET agotado_hasta = CURRENT_DATE WHERE id = :id"),
                     {"id": cid})
    cargar_componentes.clear()
    flash("Marcado agotado por hoy", "🚫")


def quitar_agotado_componente(cid: int):
    with engine.begin() as conn:
        conn.execute(text("UPDATE menu_componentes SET agotado_hasta = NULL WHERE id = :id"),
                     {"id": cid})
    cargar_componentes.clear()
    flash("Disponible de nuevo", "♻️")


# ══════════════════════════════════════════════════════════════════════════════
# DB: catálogo por categoría (tabla menu: especial / a_la_carta / bebida)
# ══════════════════════════════════════════════════════════════════════════════
def agregar_item(categoria: str, nombre: str, precio: int, descripcion=None):
    precio = int(precio)
    with engine.begin() as conn:
        mx = conn.execute(text("SELECT COALESCE(MAX(orden),0) FROM menu WHERE categoria = :c"),
                          {"c": categoria}).scalar()
        conn.execute(text(
            "INSERT INTO menu (nombre, precio, activo, orden, categoria, descripcion) "
            "VALUES (:n, :p, TRUE, :o, :c, :d)"
        ), {"n": nombre.strip(), "p": int(precio), "o": mx + 1, "c": categoria,
            "d": (descripcion or None)})
    _clear_menu_caches()
    flash("Plato agregado", "✅")


def actualizar_item(item_id: int, nombre: str, precio: int, descripcion=None):
    with engine.begin() as conn:
        conn.execute(text(
            "UPDATE menu SET nombre = :n, precio = :p, descripcion = :d WHERE id = :id"
        ), {"n": nombre.strip(), "p": int(precio), "d": (descripcion or None), "id": item_id})
    _clear_menu_caches()
    flash("Plato actualizado", "✅")


def toggle_item(item_id: int, activo_actual: bool):
    with engine.begin() as conn:
        conn.execute(text("UPDATE menu SET activo = :a WHERE id = :id"),
                     {"a": not activo_actual, "id": item_id})
    _clear_menu_caches()
    flash("Plato activado" if not activo_actual else "Plato desactivado",
          "▶️" if not activo_actual else "⏸️")


def eliminar_item(item_id: int):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM menu WHERE id = :id"), {"id": item_id})
    _clear_menu_caches()
    flash("Plato eliminado", "🗑️")


def marcar_agotado_item(item_id: int):
    with engine.begin() as conn:
        conn.execute(text("UPDATE menu SET agotado_hasta = CURRENT_DATE WHERE id = :id"),
                     {"id": item_id})
    _clear_menu_caches()
    flash("Marcado agotado por hoy", "🚫")


def quitar_agotado_item(item_id: int):
    with engine.begin() as conn:
        conn.execute(text("UPDATE menu SET agotado_hasta = NULL WHERE id = :id"),
                     {"id": item_id})
    _clear_menu_caches()
    flash("Disponible de nuevo", "♻️")


# ── Edición de stock por ítem (desde la tarjeta del menú, no solo en 📦 Inventario) ──
# El indicador de stock de cada tarjeta es además el editor: stock None = ilimitado;
# un número = porciones de hoy (0 = agotado). Invalida las cachés y refresca igual que
# el resto de acciones de la tarjeta (activar / 86 / editar).
def guardar_stock_componente(cid: int, stock):
    with engine.begin() as conn:
        conn.execute(text("UPDATE menu_componentes SET stock = :s WHERE id = :id"),
                     {"s": (None if stock is None else int(stock)), "id": int(cid)})
    cargar_componentes.clear()
    flash("Stock actualizado", "📦")


def guardar_stock_item(item_id: int, stock):
    with engine.begin() as conn:
        conn.execute(text("UPDATE menu SET stock = :s WHERE id = :id"),
                     {"s": (None if stock is None else int(stock)), "id": int(item_id)})
    _clear_menu_caches()
    flash("Stock actualizado", "📦")


# ══════════════════════════════════════════════════════════════════════════════
# DB: ajustes (precios planos + recargo de entrega)
# ══════════════════════════════════════════════════════════════════════════════
def guardar_ajustes(d: dict):
    with engine.begin() as conn:
        for clave, valor in d.items():
            conn.execute(text(
                "INSERT INTO ajustes (clave, valor) VALUES (:k, :v) "
                "ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor"
            ), {"k": clave, "v": str(valor)})
    cargar_ajustes.clear()
    _clear_menu_caches()
    flash("Ajustes guardados", "✅")


# ══════════════════════════════════════════════════════════════════════════════
# Modales (Plato del Día: renombrar / eliminar opción)
# ══════════════════════════════════════════════════════════════════════════════
@st.dialog("Editar opción")
def _dialog_editar_componente(cid: int, nombre_actual: str):
    nuevo = st.text_input("Nombre", value=nombre_actual)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("💾 Guardar", key=f"btn_guardar_comp_{cid}", type="primary",
                     use_container_width=True):
            if nuevo.strip():
                actualizar_componente(cid, nuevo)
                st.rerun()
            else:
                st.error("El nombre no puede estar vacío.")
    with c2:
        if st.button("Volver", key=f"volver_comp_{cid}", use_container_width=True):
            st.rerun()


@st.dialog("Eliminar opción")
def _dialog_eliminar_componente(cid: int, nombre: str):
    st.markdown(f"¿Eliminar **{html.escape(str(nombre))}** permanentemente?  \n"
                "Esta acción no se puede deshacer.")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("🗑 Sí, eliminar", key=f"confirm_eliminar_comp_{cid}", type="primary",
                     use_container_width=True):
            eliminar_componente(cid)
            st.rerun()
    with c2:
        if st.button("Volver", key=f"volver_del_comp_{cid}", use_container_width=True):
            st.rerun()


@st.dialog("Eliminar plato")
def _dialog_eliminar_item(item_id: int, nombre: str, categoria: str):
    st.markdown(f"¿Eliminar **{html.escape(str(nombre))}** permanentemente?  \n"
                "Esta acción no se puede deshacer.")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("🗑 Sí, eliminar", key=f"confirm_eliminar_{categoria}_{item_id}",
                     type="primary", use_container_width=True):
            eliminar_item(item_id)
            if st.session_state.get(f"ed_{categoria}_id") == item_id:
                _clear_edit(categoria)
            st.rerun()
    with c2:
        if st.button("Volver", key=f"volver_del_{categoria}_{item_id}", use_container_width=True):
            st.rerun()


def _stock_btn_label(stock) -> str:
    """Etiqueta del botón-indicador de stock: '📦 N' (con control) o '♾️' (ilimitado)."""
    s = stock_int(stock)
    return "♾️" if s is None else f"📦 {s}"


@st.dialog("📦 Stock del día")
def _dialog_stock(scope: str, oid: int, nombre: str, stock_actual):
    """Editor de stock por ítem, abierto desde su tarjeta (sustituye al viejo 86). El stock
    sustituye al 86: poner 0 = agotado (se atenúa en el menú y se oculta/bloquea al pedir);
    'Ilimitado' = sin control. scope = 'comp' (componente) | 'menu' (plato a la carta)."""
    s = stock_int(stock_actual)
    st.markdown(f"Existencias de **{html.escape(str(nombre))}**")
    ilimitado = st.checkbox("Ilimitado (sin control de stock)", value=(s is None),
                            key=f"dlgstk_unl_{scope}_{oid}")
    cantidad = st.number_input("Porciones disponibles hoy", min_value=0, max_value=_INT32_MAX,
                               step=1, value=int(s) if s is not None else 0,
                               key=f"dlgstk_qty_{scope}_{oid}", disabled=ilimitado)
    st.caption("0 = agotado. Se descuenta al crear el pedido y se reintegra al cancelar "
               "antes de 'listo'. El montaje masivo de la mañana está en 📦 Inventario.")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("💾 Guardar", key=f"dlgstk_save_{scope}_{oid}", type="primary",
                     use_container_width=True):
            val = None if ilimitado else int(cantidad or 0)
            (guardar_stock_componente if scope == "comp" else guardar_stock_item)(oid, val)
            st.rerun()
    with c2:
        if st.button("Volver", key=f"dlgstk_back_{scope}_{oid}", use_container_width=True):
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Pestaña 1 · Plato del Día (componentes por grupo)
# ══════════════════════════════════════════════════════════════════════════════
def _regla_grupo(g: dict) -> str:
    """Texto corto de la regla de selección de un grupo: 'elige 1', 'elige 3',
    'elige 2-4', 'opcional'…"""
    mn, mx = g["min_sel"], g["max_sel"]
    if mn == 0:
        return "opcional" if mx == 1 else f"opcional (hasta {mx})"
    if mn == mx:
        return f"elige {mx}"
    return f"elige {mn}-{mx}"


def _render_plato_dia():
    from db import precio_plato_dia
    grupos = cargar_grupos_pd(solo_activos=False)
    reglas = " · ".join(f"{g['etiqueta']}: {_regla_grupo(g)}"
                        for g in grupos if g.get("activo"))
    st.markdown(
        f'<div style="background:#fafaf8; border:1px solid #ececec; border-radius:10px; '
        f'padding:0.8rem 1rem; font-size:0.85rem; color:#45443e; margin-bottom:1rem;">'
        f'💡 Precio plano del Plato del Día: <b>${fmt_money(precio_plato_dia())}</b> '
        f'(se cambia en ⚙️ Ajustes) · {html.escape(reglas)}<br>'
        f'Los pasos de selección se configuran abajo en 🧩 Grupos.</div>',
        unsafe_allow_html=True,
    )

    # Cada grupo es un acordeón a lo ANCHO en vez de una columna alta: las opciones
    # crecían mucho hacia abajo. Solo se despliega el grupo que el admin está editando;
    # el primero queda abierto para orientar. El estado abierto/cerrado sobrevive a los
    # st.rerun() de cada acción (mismo mecanismo _acc_header que los acordeones del
    # catálogo). Los grupos ya no son fijos: salen de la tabla plato_dia_grupos.
    df = cargar_componentes()
    for idx, g in enumerate(grupos):
        sub = df[df["grupo"] == g["clave"]] if not df.empty else df
        _render_grupo(sub, g, default_open=(idx == 0))

    # Componentes huérfanos: su 'grupo' no coincide con ningún grupo definido (p. ej.
    # el grupo se borró o la clave cambió a mano en la BD). Se muestran para que no
    # desaparezcan en silencio.
    if not df.empty:
        claves = {g["clave"] for g in grupos}
        huerf = df[~df["grupo"].isin(claves)]
        if not huerf.empty:
            st.warning(f"⚠️ Hay {len(huerf)} opción(es) cuyo grupo ya no existe: "
                       + ", ".join(f"{r['nombre']} ({r['grupo']})"
                                   for _, r in huerf.iterrows()))

    _render_grupos_editor(grupos)


def _render_grupo(sub, g: dict, *, default_open: bool = False):
    grupo = g["clave"]
    label = g["etiqueta"]
    activos = int((sub["activo"] == True).sum()) if not sub.empty else 0
    titulo = f"{label} · {activos} activo(s) · {_regla_grupo(g)}"
    if not g.get("activo"):
        titulo += " · ⏸ grupo desactivado"
    if not _acc_header(f"grupo_{grupo}", titulo, default_open=default_open):
        return

    if grupo == "entrada":
        st.caption("Las sopas activas aparecen aquí junto a Fruta y Huevo. "
                   "Desactiva o marca 86 una sopa para ocultarla del día.")

    if sub.empty:
        st.markdown('<p style="color:#a3a39b; font-size:0.82rem;">Sin opciones.</p>',
                    unsafe_allow_html=True)
    else:
        # Buscador PROPIO de este grupo: filtra sus opciones por nombre (útil cuando el
        # grupo tiene muchas: p. ej. decenas de proteínas o acompañamientos).
        sub, buscando = _filtrar_seccion(sub, f"buscar_comp_{grupo}", f"Buscar en {label}…")
        # Disponibles primero; inactivos / agotados / 86 bajan al fondo del grupo (mismo
        # criterio que las categorías del catálogo). Dentro del acordeón las tarjetas van
        # en una rejilla de 3 para no volver a alargar la pantalla.
        disp, no = _partir_disponibles(sub)
        cards = disp + no
        if not cards:
            st.markdown('<p style="color:#a3a39b; font-size:0.82rem;">Sin coincidencias.</p>'
                        if buscando else
                        '<p style="color:#a3a39b; font-size:0.82rem;">Sin opciones.</p>',
                        unsafe_allow_html=True)
        ncols = 3
        for i in range(0, len(cards), ncols):
            fila = st.columns(ncols)
            for j, col in enumerate(fila):
                if i + j < len(cards):
                    with col:
                        _componente_card(cards[i + j], grupo)

    # Alta rápida de una opción (la clave se rota con un nonce para limpiar el input).
    # En una columna estrecha para no estirar el botón a todo el ancho del acordeón.
    nonce = st.session_state.get(f"nonce_add_{grupo}", 0)
    with st.columns([1, 2])[0]:
        nuevo = st.text_input("Agregar", key=f"add_comp_{grupo}_{nonce}",
                              placeholder="Nueva opción…", label_visibility="collapsed")
        if st.button("➕ Agregar", key=f"btn_add_comp_{grupo}", use_container_width=True):
            if nuevo.strip():
                agregar_componente(grupo, nuevo)
                st.session_state[f"nonce_add_{grupo}"] = nonce + 1
                st.rerun()
            else:
                st.warning("Escribe un nombre.")


def _componente_card(row, grupo: str):
    cid = int(row["id"])
    nombre = row["nombre"]
    activo = bool(row["activo"])
    # "Agotado" se muestra por dos vías: 86 manual del día (agotado_hasta) o stock en 0.
    # El badge/atenuado refleja ambas; el botón 86/♻ (b2) solo gestiona el 86 MANUAL
    # (el ♻ solo deshace eso; el stock se ajusta en 📦 Inventario).
    agotado_dia = _agotado_hoy(row.get("agotado_hasta"))
    sin_stock = agotado_por_stock(row.get("stock"))
    card_cls = "menu-card" if (activo and not (agotado_dia or sin_stock)) else "menu-card inactivo"
    if agotado_dia:
        estado = '<span class="badge badge-agotado">🚫 Hoy</span>'
    elif sin_stock:
        estado = '<span class="badge badge-agotado">🚫 Sin stock</span>'
    elif activo:
        estado = '<span class="badge badge-activo">●</span>'
    else:
        estado = '<span class="badge badge-inactivo">○</span>'

    st.markdown(
        f'<div class="{card_cls}" style="padding:0.6rem 0.8rem; margin-bottom:0.4rem;">'
        f'<div style="display:flex; justify-content:space-between; align-items:center;">'
        f'<div class="menu-nombre" style="font-size:0.9rem;">{html.escape(str(nombre))}</div>'
        f'<div>{estado}</div></div></div>',
        unsafe_allow_html=True,
    )
    b1, b2, b3, b4 = st.columns(4)
    with b1:
        if st.button("⏸" if activo else "▶", key=f"toggle_comp_{cid}",
                     help="Desactivar" if activo else "Activar", use_container_width=True):
            toggle_componente(cid, activo)
            st.rerun()
    with b2:
        # El indicador de stock ocupa el lugar del antiguo 86; al pulsarlo se edita.
        if st.button(_stock_btn_label(row.get("stock")), key=f"stock_comp_{cid}",
                     help="Stock de hoy (toca para editar). 0 = agotado.",
                     use_container_width=True):
            _dialog_stock("comp", cid, str(nombre), row.get("stock"))
    with b3:
        if st.button("✏️", key=f"edit_comp_{cid}", help="Renombrar", use_container_width=True):
            _dialog_editar_componente(cid, str(nombre))
    with b4:
        if st.button("🗑", key=f"eliminar_comp_{cid}", help="Eliminar", use_container_width=True):
            _dialog_eliminar_componente(cid, str(nombre))


# ══════════════════════════════════════════════════════════════════════════════
# Editor de GRUPOS del Plato del Día (pasos de selección, tabla plato_dia_grupos)
# ══════════════════════════════════════════════════════════════════════════════
def _render_grupos_editor(grupos):
    """Gestión de los pasos de selección: etiqueta, orden, mín/máx, repetir, activo,
    alta de grupos nuevos y borrado (solo si el grupo no tiene opciones cargadas)."""
    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
    if not _acc_header("grupos_editor", "🧩 Grupos (pasos de selección del Plato del Día)",
                       default_open=False):
        return
    # id=0 = fallback en memoria (tabla sin sembrar): nada que editar todavía.
    if any(int(g.get("id") or 0) == 0 for g in grupos):
        st.info("Los grupos aún no están sembrados en la base de datos; se usan los "
                "clásicos por defecto. Arranca el bot (o corre el script de "
                "aprovisionamiento) y vuelve aquí para editarlos.")
        return
    st.caption("Cada grupo es un paso del configurador del Plato del Día. "
               "Mín 0 = opcional · Máx = cuántas opciones elige el cliente · "
               "Repetir = puede llevar 2x la misma opción. La clave (entre paréntesis) "
               "no se edita: ancla las opciones ya cargadas del grupo.")
    h = st.columns([2.4, 0.9, 0.9, 0.9, 1.0, 1.0, 0.7])
    for col, t in zip(h, ["Etiqueta", "Orden", "Mín", "Máx", "Repetir", "Activo", ""]):
        col.markdown(f"<span style='font-size:0.72rem; color:#a3a39b;'>{t}</span>",
                     unsafe_allow_html=True)
    vals = {}
    for g in grupos:
        gid = int(g["id"])
        c1, c2, c3, c4, c5, c6, c7 = st.columns([2.4, 0.9, 0.9, 0.9, 1.0, 1.0, 0.7])
        with c1:
            et = st.text_input("Etiqueta", value=g["etiqueta"], key=f"gpd_et_{gid}",
                               label_visibility="collapsed")
            st.caption(f"({g['clave']})")
        with c2:
            orden = st.number_input("Orden", min_value=0, step=1, value=int(g["orden"]),
                                    key=f"gpd_or_{gid}", label_visibility="collapsed")
        with c3:
            mn = st.number_input("Mín", min_value=0, max_value=9, step=1,
                                 value=int(g["min_sel"]), key=f"gpd_mn_{gid}",
                                 label_visibility="collapsed")
        with c4:
            mx = st.number_input("Máx", min_value=1, max_value=9, step=1,
                                 value=int(g["max_sel"]), key=f"gpd_mx_{gid}",
                                 label_visibility="collapsed")
        with c5:
            rep = st.checkbox("Repetir", value=bool(g["permite_repetir"]),
                              key=f"gpd_rep_{gid}")
        with c6:
            act = st.checkbox("Activo", value=bool(g["activo"]), key=f"gpd_act_{gid}")
        with c7:
            if st.button("🗑", key=f"gpd_del_{gid}",
                         help="Eliminar grupo (solo si no tiene opciones cargadas)"):
                err = eliminar_grupo_pd(gid)
                flash(err, "⚠️") if err else flash("Grupo eliminado", "🗑")
                st.rerun()
        vals[gid] = (et, orden, mn, mx, rep, act)

    if st.button("💾 Guardar grupos", type="primary", key="gpd_guardar"):
        for gid, (et, orden, mn, mx, rep, act) in vals.items():
            guardar_grupo_pd(gid, etiqueta=et, orden=int(orden), min_sel=int(mn),
                             max_sel=int(mx), permite_repetir=bool(rep),
                             activo=bool(act))
        flash("Grupos guardados", "🧩")
        st.rerun()

    # Alta de un grupo nuevo (p. ej. 'Sopa' separada de la entrada, 'Postre'…).
    st.markdown(titulo_seccion("➕ Nuevo grupo", style="margin-top:0.8rem;"),
                unsafe_allow_html=True)
    nonce = st.session_state.get("gpd_new_nonce", 0)
    c1, c2, c3, c4 = st.columns([2.4, 0.9, 0.9, 1.6])
    with c1:
        net = st.text_input("Etiqueta del grupo", key=f"gpd_new_et_{nonce}",
                            placeholder="P. ej. Sopa, Postre…")
    with c2:
        nmn = st.number_input("Mín", min_value=0, max_value=9, step=1, value=1,
                              key=f"gpd_new_mn_{nonce}")
    with c3:
        nmx = st.number_input("Máx", min_value=1, max_value=9, step=1, value=1,
                              key=f"gpd_new_mx_{nonce}")
    with c4:
        nrep = st.checkbox("Puede repetir la misma opción", key=f"gpd_new_rep_{nonce}")
    if st.button("➕ Crear grupo", key="gpd_new_btn"):
        err = crear_grupo_pd(net, net, min_sel=int(nmn), max_sel=int(nmx),
                             permite_repetir=bool(nrep))
        if err:
            st.error(err)
        else:
            st.session_state["gpd_new_nonce"] = nonce + 1
            flash("Grupo creado", "🧩")
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Editor de CATEGORÍAS del catálogo (pestañas del catálogo, tabla categorias)
# ══════════════════════════════════════════════════════════════════════════════
def _render_categorias_editor():
    """Gestión de las categorías del catálogo: nombre, emoji, orden, activa/inactiva,
    horario opcional (SOLO oculta en la carta digital del cliente), extras incluidos
    (qué grupos del Plato del Día vienen sin costo, ver db.extras_de_categoria) y
    borrado (solo si la categoría no tiene platos cargados). Alta de categorías nuevas
    al final. El comportamiento fijo (descripción / recargo de entrega) NO se edita
    aquí a propósito: una categoría nueva se comporta como 'A la carta'."""
    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
    if not _acc_header("categorias_editor", "🏷️ Categorías (pestañas del catálogo)",
                       default_open=False):
        return
    categorias = cargar_categorias(solo_activas=False)
    # id=0 = fallback en memoria (tabla sin sembrar): nada que editar todavía.
    if any(int(c.get("id") or 0) == 0 for c in categorias):
        st.info("Las categorías aún no están sembradas en la base de datos; se usan las "
                "4 clásicas por defecto. Arranca el bot (o corre el script de "
                "aprovisionamiento) y vuelve aquí para editarlas.")
        return
    st.caption("Cada categoría es una pestaña del catálogo (aquí, en el POS del mesero y "
               "en la carta digital). Una categoría nueva se agrega con el mismo "
               "comportamiento de 'A la carta' y SIN extras incluidos. El horario es "
               "opcional y SOLO oculta la categoría en la carta digital del cliente "
               "fuera de rango — el panel y el POS siempre la ven, sin importar la "
               "hora, y el cambio tarda hasta un minuto en verse en la carta ya abierta "
               "(se refresca sola). La clave (entre paréntesis) no se edita: ancla los "
               "platos ya cargados.")
    h = st.columns([0.8, 2.2, 0.8, 0.9, 0.7])
    for col, t in zip(h, ["Emoji", "Nombre", "Orden", "Activa", ""]):
        col.markdown(f"<span style='font-size:0.72rem; color:#a3a39b;'>{t}</span>",
                     unsafe_allow_html=True)
    # Hora del negocio para el aviso 'visible/oculta ahora' de cada horario. Una sola vez
    # para todas las filas: así el editor entero razona contra el mismo instante.
    ahora = ahora_bogota().time()
    # Grupos del Plato del Día ofrecibles como extra: solo los de selección ÚNICA
    # (max_sel==1) — el selector de extras es un radio de una opción, no soporta
    # grupos multi como acompañamientos. {clave: etiqueta}, en su 'orden'.
    opciones_extra = {g["clave"]: g["etiqueta"] for g in cargar_grupos_pd()
                      if int(g.get("max_sel") or 1) == 1}
    comp_activos = componentes_activos_por_grupo()
    vals = {}
    for c in categorias:
        cid = int(c["id"])
        c1, c2, c3, c4, c5 = st.columns([0.8, 2.2, 0.8, 0.9, 0.7])
        with c1:
            emoji = st.text_input("Emoji", value=c.get("emoji") or "", key=f"cat_em_{cid}",
                                  label_visibility="collapsed", max_chars=8)
        with c2:
            et = st.text_input("Nombre", value=c["etiqueta"], key=f"cat_et_{cid}",
                               label_visibility="collapsed")
            st.caption(f"({c['clave']})")
        with c3:
            orden = st.number_input("Orden", min_value=0, step=1, value=int(c["orden"]),
                                    key=f"cat_or_{cid}", label_visibility="collapsed")
        with c4:
            act = st.checkbox("Activa", value=bool(c["activo"]), key=f"cat_act_{cid}")
        with c5:
            if st.button("🗑", key=f"cat_del_{cid}",
                         help="Eliminar categoría (solo si no tiene platos cargados)"):
                err = eliminar_categoria(cid)
                flash(err, "⚠️") if err else flash("Categoría eliminada", "🗑")
                st.rerun()

        con_horario = st.checkbox(
            "🕐 Restringir horario (se oculta en la carta digital fuera de rango)",
            value=(c.get("disponible_desde") is not None
                   or c.get("disponible_hasta") is not None),
            key=f"cat_hor_on_{cid}",
        )
        desde = hasta = None
        if con_horario:
            hc1, hc2 = st.columns(2)
            with hc1:
                desde = st.time_input("Desde", value=c.get("disponible_desde") or dtime(6, 0),
                                      key=f"cat_desde_{cid}")
            with hc2:
                hasta = st.time_input("Hasta", value=c.get("disponible_hasta") or dtime(11, 0),
                                      key=f"cat_hasta_{cid}")
            # Efecto del horario AHORA MISMO, con los valores tecleados (aún sin guardar).
            # Responde de una a "puse el horario y no pasa nada": aquí se ve si la carta
            # del cliente la está mostrando, sin tener que abrirla en el celular. 'Desde'
            # mayor que 'Hasta' es una franja nocturna válida (p. ej. 20:00–02:00).
            if not act:
                chip = '<span class="badge badge-inactivo">○ Oculta · categoría desactivada</span>'
            elif en_horario_categoria({"disponible_desde": desde,
                                       "disponible_hasta": hasta}, ahora):
                chip = '<span class="badge badge-activo">● Visible ahora en la carta</span>'
            else:
                chip = '<span class="badge badge-inactivo">○ Oculta ahora en la carta</span>'
            st.markdown(f'{chip} <span style="font-size:0.72rem; color:#a3a39b;">'
                        f'son las {ahora.strftime("%H:%M")} en Bogotá</span>',
                        unsafe_allow_html=True)

        extras_actuales = extras_de_categoria(c)
        ex_on = st.checkbox("🎁 Incluye entrada/bebida sin costo", value=bool(extras_actuales),
                            key=f"cat_ex_on_{cid}")
        extras_sel = []
        if ex_on:
            if opciones_extra:
                default = [g for g in extras_actuales if g in opciones_extra]
                extras_sel = st.multiselect(
                    "Grupos incluidos", options=list(opciones_extra.keys()), default=default,
                    key=f"cat_ex_grupos_{cid}", format_func=lambda g: opciones_extra.get(g, g),
                    label_visibility="collapsed")
                if extras_sel and not any(comp_activos.get(g) for g in extras_sel):
                    st.caption("⚠️ Ninguna opción activa en esos grupos todavía: no se "
                               "mostrará nada hasta que actives alguna en 🍽️ Plato del Día.")
            else:
                st.caption("⚠️ No hay grupos de selección única en el Plato del Día "
                           "(ver 🍽️ Plato del Día → 🧩 Grupos).")
        st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
        vals[cid] = (et, emoji, orden, act, desde, hasta, extras_sel)

    if st.button("💾 Guardar categorías", type="primary", key="cat_guardar"):
        for cid, (et, emoji, orden, act, desde, hasta, extras_sel) in vals.items():
            guardar_categoria(cid, etiqueta=et, emoji=emoji, orden=int(orden), activo=bool(act),
                              disponible_desde=desde, disponible_hasta=hasta,
                              extras_grupos=extras_sel)
        flash("Categorías guardadas", "🏷️")
        st.rerun()

    # Alta de una categoría nueva (p. ej. Desayunos, Postres…).
    st.markdown(titulo_seccion("➕ Nueva categoría", style="margin-top:0.8rem;"),
                unsafe_allow_html=True)
    nonce = st.session_state.get("cat_new_nonce", 0)
    c1, c2 = st.columns([0.8, 2.2])
    with c1:
        nem = st.text_input("Emoji", key=f"cat_new_em_{nonce}", placeholder="🍳", max_chars=8)
    with c2:
        net = st.text_input("Nombre de la categoría", key=f"cat_new_et_{nonce}",
                            placeholder="P. ej. Desayunos, Postres…")
    if st.button("➕ Crear categoría", key="cat_new_btn"):
        err = crear_categoria(net, net, emoji=nem)
        if err:
            st.error(err)
        else:
            st.session_state["cat_new_nonce"] = nonce + 1
            flash("Categoría creada", "🏷️")
            # El importador reconoce cada categoría por su etiqueta (ver _seccion_map), así
            # que un nombre que YA era alias de una clásica ('Extras' → Adicionales, 'Carta'
            # → A la carta) pasa a enrutar a la nueva. Es lo correcto —lo explícito del
            # restaurante gana— pero cambia en silencio el destino de hojas de Excel que ya
            # venía usando: se avisa en vez de dejarlo pasar.
            alias = _ALIAS_MENU_CLASICOS.get(_norm(net))
            if alias:
                flash(f"Ojo: al importar, «{net.strip()}» ahora entra a esta categoría "
                      f"y ya no a «{_dest_label(alias)}».", "⚠️")
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Pestañas 2–4 · Catálogo (Especiales / A la carta / Bebidas)
# ══════════════════════════════════════════════════════════════════════════════
def _enter_edit(categoria: str, row, con_precio: bool, con_desc: bool = False):
    # Escribimos DIRECTO en las claves de los widgets (no en defaults aparte) para
    # esquivar el gotcha de Streamlit: un widget con key ignora value= si ya hay estado.
    st.session_state[f"ed_{categoria}_id"] = int(row["id"])
    st.session_state[f"in_{categoria}_nombre"] = str(row["nombre"])
    if con_precio:
        st.session_state[f"in_{categoria}_precio"] = int(row["precio"])
    if con_desc:
        st.session_state[f"in_{categoria}_desc"] = _txt(row.get("descripcion"))


def _clear_edit(categoria: str):
    for suf in ("nombre", "precio", "desc"):
        st.session_state.pop(f"in_{categoria}_{suf}", None)
    st.session_state.pop(f"ed_{categoria}_id", None)


def _render_catalogo_tab(categoria: str, label: str, con_precio: bool, con_desc: bool = False):
    df = cargar_catalogo()
    sub = df[df["categoria"] == categoria].copy() if not df.empty else df

    col_lista, col_form = st.columns([2, 1])

    with col_lista:
        st.markdown(f'<div class="section-title">{label}</div>', unsafe_allow_html=True)

        if sub is None or sub.empty:
            st.markdown('<p style="color:#a3a39b; font-size:0.85rem;">No hay platos en esta sección.</p>',
                        unsafe_allow_html=True)
        else:
            # Buscador PROPIO de esta categoría: filtra sus platos por nombre o descripción
            # (útil cuando la sección tiene muchos ítems).
            sub, buscando = _filtrar_seccion(sub, f"buscar_cat_{categoria}", f"Buscar en {label}…")
            # Los PLATOS de la categoría se agrupan en dos acordeones plegables:
            # Disponibles (activo + con stock + sin 86) abierto al entrar, y No disponibles
            # (inactivo / agotado / 86) plegado. Así se ve de inmediato lo que se puede
            # vender y lo no disponible queda recogido hasta que se necesite.
            disp, no = _partir_disponibles(sub)
            if buscando and not disp and not no:
                st.markdown('<p style="color:#a3a39b; font-size:0.82rem;">Sin coincidencias.</p>',
                            unsafe_allow_html=True)
            if _acc_header(f"disp_{categoria}", f"● Disponibles · {len(disp)}", default_open=True):
                if disp:
                    for row in disp:
                        _item_card(row, categoria, con_precio, con_desc)
                else:
                    st.markdown('<p style="color:#a3a39b; font-size:0.82rem;">Ninguno disponible ahora.</p>',
                                unsafe_allow_html=True)
            if no:
                if _acc_header(f"nodisp_{categoria}", f"○ No disponibles · {len(no)}", default_open=False):
                    for row in no:
                        _item_card(row, categoria, con_precio, con_desc)

    with col_form:
        _item_form(categoria, label, con_precio, con_desc)


def _item_card(row, categoria: str, con_precio: bool, con_desc: bool = False):
    iid = int(row["id"])
    nombre = row["nombre"]
    precio = int(row["precio"])
    activo = bool(row["activo"])
    # Agotado por 86 manual (agotado_hasta) o por stock en 0. El badge/atenuado refleja
    # ambas; el botón 86/♻ (b2) gestiona solo el 86 manual (el stock va en 📦 Inventario).
    agotado_dia = _agotado_hoy(row.get("agotado_hasta"))
    sin_stock = agotado_por_stock(row.get("stock"))
    desc = _txt(row.get("descripcion"))
    card_cls = "menu-card" if (activo and not (agotado_dia or sin_stock)) else "menu-card inactivo"
    badge = ('<span class="badge badge-activo">● Activo</span>' if activo
             else '<span class="badge badge-inactivo">○ Inactivo</span>')
    if agotado_dia:
        badge += ' <span class="badge badge-agotado">🚫 Hoy</span>'
    elif sin_stock:
        badge += ' <span class="badge badge-agotado">🚫 Sin stock</span>'

    sub_parts = []
    if con_precio:
        sub_parts.append(f'<div class="menu-precio">${fmt_money(precio)}</div>')
    if con_desc:
        if desc:
            sub_parts.append(f'<div class="menu-precio" style="font-style:italic;">'
                             f'{html.escape(desc)}</div>')
        else:
            sub_parts.append('<div class="menu-precio" style="color:#d97706;">Sin descripción</div>')
    sub_html = "".join(sub_parts)

    col_card, col_btns = st.columns([3, 1.4])
    with col_card:
        st.markdown(
            f'<div class="{card_cls}">'
            f'<div style="display:flex; justify-content:space-between; align-items:center;">'
            f'<div><div class="menu-nombre">{html.escape(str(nombre))}</div>{sub_html}</div>'
            f'<div style="text-align:right;">{badge}</div></div></div>',
            unsafe_allow_html=True,
        )
    with col_btns:
        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        b1, b2, b3, b4 = st.columns(4)
        with b1:
            if st.button("⏸" if activo else "▶", key=f"toggle_{categoria}_{iid}",
                         help="Desactivar" if activo else "Activar", use_container_width=True):
                toggle_item(iid, activo)
                st.rerun()
        with b2:
            # Indicador de stock en el lugar del antiguo 86; al pulsarlo se edita.
            if st.button(_stock_btn_label(row.get("stock")), key=f"stock_{categoria}_{iid}",
                         help="Stock de hoy (toca para editar). 0 = agotado.",
                         use_container_width=True):
                _dialog_stock("menu", iid, str(nombre), row.get("stock"))
        with b3:
            if st.button("✏️", key=f"edit_{categoria}_{iid}", help="Editar",
                         use_container_width=True):
                _enter_edit(categoria, row, con_precio, con_desc)
                st.rerun()
        with b4:
            if st.button("🗑", key=f"eliminar_{categoria}_{iid}", help="Eliminar",
                         use_container_width=True):
                _dialog_eliminar_item(iid, str(nombre), categoria)


def _item_form(categoria: str, label: str, con_precio: bool, con_desc: bool = False):
    editando = f"ed_{categoria}_id" in st.session_state
    st.markdown(f'<div class="section-title">{"Editar" if editando else "Agregar"}</div>',
                unsafe_allow_html=True)

    nombre = st.text_input("Nombre del plato", key=f"in_{categoria}_nombre")

    precio_val = 0
    if con_precio:
        precio_val = st.number_input("Precio", min_value=0, max_value=_INT32_MAX, step=1000,
                                     key=f"in_{categoria}_precio")

    desc = None
    if con_desc:
        desc = st.text_area("Descripción (qué incluye)", key=f"in_{categoria}_desc",
                            placeholder="Ej: incluye todos los acompañamientos")

    b_guardar, b_cancelar = st.columns(2)
    with b_guardar:
        if st.button("💾 Guardar", type="primary", key=f"btn_guardar_{categoria}",
                     use_container_width=True):
            if not (nombre or "").strip():
                st.error("El nombre no puede estar vacío.")
            elif con_precio and int(precio_val or 0) <= 0:
                st.error("El precio debe ser mayor a 0.")
            else:
                if editando:
                    actualizar_item(
                        st.session_state[f"ed_{categoria}_id"], nombre, precio_val, desc,
                    )
                    _clear_edit(categoria)
                else:
                    agregar_item(categoria, nombre, precio_val, desc)
                    _clear_edit(categoria)
                st.rerun()
    with b_cancelar:
        if editando:
            if st.button("✕ Cancelar", key=f"btn_cancelar_{categoria}", use_container_width=True):
                _clear_edit(categoria)
                st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Pestaña 5 · Ajustes (precios planos + recargo de entrega)
# ══════════════════════════════════════════════════════════════════════════════
def _int_aj(aj: dict, clave: str, default: int) -> int:
    try:
        return int(float(aj.get(clave, default)))
    except (TypeError, ValueError):
        return default


def _render_ajustes():
    from db import metodos_pago
    aj = cargar_ajustes()
    st.markdown('<div class="section-title">Precios y recargos</div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        pd_precio = st.number_input("Precio Plato del Día", min_value=0, step=1000,
                                    value=_int_aj(aj, "plato_dia_precio", 0), key="aj_plato_dia")
    with c2:
        fee = st.number_input("Recargo de entrega (Domicilio / Para Llevar)", min_value=0,
                              step=500, value=_int_aj(aj, "fee_entrega", 0), key="aj_fee")
    st.caption("El nº de acompañamientos (y todos los pasos del Plato del Día) se "
               "configura en 🍽️ Plato del Día → 🧩 Grupos.")

    # ── Identidad del restaurante (antes quemada en el código) ────────────────
    st.markdown('<div class="section-title" style="margin-top:1rem;">Identidad del '
                'restaurante</div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1.6, 1.2, 1])
    with c1:
        r_nombre = st.text_input("Nombre", value=str(aj.get("restaurante_nombre") or ""),
                                 key="aj_r_nombre",
                                 help="Aparece en el panel, el saludo del bot y los recibos.")
    with c2:
        r_dir = st.text_input("Dirección", value=str(aj.get("restaurante_direccion") or ""),
                              key="aj_r_dir", help="Se imprime en el encabezado del recibo.")
    with c3:
        r_tel = st.text_input("Teléfono", value=str(aj.get("restaurante_telefono") or ""),
                              key="aj_r_tel", help="Se imprime en el encabezado del recibo.")
    saludo = st.text_area(
        "Saludo del bot de WhatsApp", value=str(aj.get("bot_saludo") or ""),
        key="aj_saludo", height=140,
        help="Plantilla del mensaje de bienvenida. {nombre} = nombre del restaurante, "
             "{link} = enlace a la carta digital. Déjalo vacío para usar el texto por defecto.")

    # ── Métodos de pago (antes quemados: Nequi/Daviplata/Bre-B) ───────────────
    st.markdown('<div class="section-title" style="margin-top:1rem;">Métodos de pago</div>',
                unsafe_allow_html=True)
    mp = metodos_pago()
    c1, c2 = st.columns([1, 2])
    with c1:
        mp_ef = st.checkbox("Acepta efectivo", value=bool(mp.get("efectivo", True)),
                            key="aj_mp_ef")
    with c2:
        mp_txt = st.text_area(
            "Transferencias aceptadas (una por línea: clave=Etiqueta, o solo la etiqueta)",
            value="\n".join(f"{k}={v}" for k, v in mp.get("transferencia", {}).items()),
            key="aj_mp_txt", height=100,
            help="Ejemplos: nequi=Nequi · daviplata=Daviplata · breb=Bre-B. "
                 "Deja el cuadro vacío si el restaurante no recibe transferencias.")

    if st.button("💾 Guardar ajustes", type="primary", key="btn_guardar_ajustes"):
        transf = {}
        for linea in (mp_txt or "").splitlines():
            linea = linea.strip()
            if not linea:
                continue
            k, _, v = linea.partition("=")
            clave_mp = re.sub(r"[^a-z0-9_]", "", k.strip().lower()
                              .replace(" ", "_").replace("ñ", "n"))[:20]
            if clave_mp:
                transf[clave_mp] = v.strip() or k.strip()
        guardar_ajustes({
            "plato_dia_precio": int(pd_precio),
            "fee_entrega": int(fee),
            "restaurante_nombre": r_nombre.strip(),
            "restaurante_direccion": r_dir.strip(),
            "restaurante_telefono": r_tel.strip(),
            "bot_saludo": saludo.strip(),
            "metodos_pago": json.dumps(
                {"efectivo": bool(mp_ef), "transferencia": transf}, ensure_ascii=False),
        })
        st.rerun()

    st.markdown(
        '<div style="background:#fafaf8; border:1px solid #ececec; border-radius:10px; '
        'padding:1rem; font-size:0.78rem; color:#6b6b64; margin-top:1rem;">'
        '<div style="color:#45443e; font-weight:600; margin-bottom:6px;">💡 Cómo se aplican</div>'
        'El <b>Plato del Día</b> cuesta lo mismo sin importar la combinación elegida.<br>'
        'Cada <b>Especial</b> tiene su propio precio (se edita en la pestaña ⭐ Especiales).<br>'
        'El <b>recargo de entrega</b> se suma una vez a cada pedido de Domicilio o Para Llevar.<br>'
        'El <b>nombre</b> sale en el panel, el bot y los recibos; el <b>saludo</b> lo envía el '
        'bot al primer mensaje; los <b>métodos de pago</b> alimentan el cobro en caja y la app '
        'del cliente.</div>',
        unsafe_allow_html=True,
    )

    _render_categorias_editor()


# ══════════════════════════════════════════════════════════════════════════════
# Pestaña 6 · Inventario del día (stock por componente y por plato)
# ══════════════════════════════════════════════════════════════════════════════
# Pantalla de montaje diario: cada mañana el administrador fija cuántas porciones hay
# de cada micro-componente del Plato del Día (cada sopa, proteína, acompañamiento…) y
# cuántas unidades de cada plato a la carta / especial / bebida. Al guardar se inicializan
# o sobrescriben los contadores en vivo. Marcar "Ilimitado" deja el ítem SIN control
# (no se descuenta ni se oculta). Va en un st.form para no recargar en cada tecla.
def _actual_txt(stock) -> str:
    """Texto del stock vivo de una fila ('actual: N' / 'actual: ilimitado')."""
    s = stock_int(stock)
    return "actual: ilimitado" if s is None else f"actual: {s}"


def _fila_inventario(scope: str, oid: int, nombre, stock_actual):
    """Una fila del formulario: nombre + 'Ilimitado' + 'Cantidad'. Devuelve el par
    (ilimitado, cantidad) tal como quedan los widgets (se lee en el submit)."""
    s = stock_int(stock_actual)
    c_n, c_u, c_q = st.columns([3, 1.4, 1.6])
    with c_n:
        st.markdown(
            f'<div style="padding:6px 0;"><span style="font-size:0.9rem; color:#26262b;">'
            f'{html.escape(str(nombre))}</span>'
            f'<div style="font-size:0.72rem; color:#a3a39b;">{_actual_txt(stock_actual)}</div></div>',
            unsafe_allow_html=True,
        )
    with c_u:
        ilimitado = st.checkbox("Ilimitado", value=(s is None),
                                key=f"inv_{scope}_unl_{oid}")
    with c_q:
        cantidad = st.number_input("Cantidad", min_value=0, max_value=_INT32_MAX, step=1,
                                   value=int(s) if s is not None else 0,
                                   key=f"inv_{scope}_qty_{oid}", label_visibility="collapsed")
    return ilimitado, int(cantidad or 0)


def _resumen_inv(sub) -> str:
    """Resumen para el título de un acordeón (se ve sin desplegar): cuántas opciones
    llevan control y cuántas están agotadas. Ej: '2/3 con control · 1 agotado'."""
    total = len(sub)
    con = sum(1 for _, r in sub.iterrows() if stock_int(r.get("stock")) is not None)
    agot = sum(1 for _, r in sub.iterrows() if agotado_por_stock(r.get("stock")))
    txt = f"{con}/{total} con control"
    if agot:
        txt += f" · {agot} agotado"
    return txt


def _render_inventario():
    st.markdown(
        '<div style="background:#fafaf8; border:1px solid #ececec; border-radius:10px; '
        'padding:0.8rem 1rem; font-size:0.85rem; color:#45443e; margin-bottom:1rem;">'
        '📦 Fija el <b>stock del día</b>. Cada componente del Plato del Día se cuenta por '
        'separado (50 sopas, 50 ensaladas, 50 res, 50 pollo…) y cada plato a la carta como '
        'una unidad. Al guardar se <b>sobrescriben</b> los contadores en vivo. Marca '
        '<b>Ilimitado</b> para no controlar un ítem.</div>',
        unsafe_allow_html=True,
    )

    df_comp = cargar_componentes()
    df_cat = cargar_catalogo()
    comp_vals, menu_vals = {}, {}

    with st.form("form_inventario_dia"):
        st.markdown(titulo_seccion('🍛 Plato del Día · por componente'),
                    unsafe_allow_html=True)
        if df_comp is None or df_comp.empty:
            st.caption("No hay componentes del Plato del Día. Créalos en la pestaña 🍽️ Plato del Día.")
        else:
            # Un acordeón por grupo (colapsado): mantiene la pantalla compacta y solo se
            # despliega hacia abajo el grupo que el admin toca. El título lleva el resumen.
            etiquetas = etiquetas_grupos_pd()
            for grupo in list(dict.fromkeys(
                    [g["clave"] for g in cargar_grupos_pd(solo_activos=False)]
                    + df_comp["grupo"].tolist())):
                sub = df_comp[df_comp["grupo"] == grupo]
                if sub.empty:
                    continue
                with st.expander(f"{etiquetas.get(grupo, grupo)} · {_resumen_inv(sub)}",
                                 expanded=False):
                    for _, row in sub.iterrows():
                        cid = int(row["id"])
                        comp_vals[cid] = _fila_inventario("comp", cid, row["nombre"], row.get("stock"))

        st.markdown(titulo_seccion('🍽️ Platos a la carta · por unidad', style="margin-top:1rem;"),
                    unsafe_allow_html=True)
        # Un acordeón por categoría VIVA del catálogo (las clásicas y las que el
        # restaurante creó en ⚙️ Ajustes → 🏷️ Categorías), más las categorías huérfanas
        # que aún tengan platos: el montaje de la mañana no puede dejar ítems fuera. Antes
        # esta lista estaba quemada a especial/a_la_carta/bebida y se tragaba en silencio
        # los Adicionales y TODA categoría nueva. Mismo criterio que los grupos de arriba.
        cats = cargar_categorias(solo_activas=False)
        etiquetas_cat = {c["clave"]: f'{c["emoji"]} {c["etiqueta"]}'.strip() for c in cats}
        hay_cat = False
        for categoria in dict.fromkeys(
                [c["clave"] for c in cats]
                + (df_cat["categoria"].tolist() if not df_cat.empty else [])):
            sub = df_cat[df_cat["categoria"] == categoria] if not df_cat.empty else df_cat
            if sub is None or sub.empty:
                continue
            hay_cat = True
            with st.expander(f"{etiquetas_cat.get(categoria, categoria)} · {_resumen_inv(sub)}",
                             expanded=False):
                for _, row in sub.iterrows():
                    mid = int(row["id"])
                    menu_vals[mid] = _fila_inventario("menu", mid, row["nombre"], row.get("stock"))
        if not hay_cat:
            st.caption("No hay platos a la carta. Agrégalos en sus pestañas.")

        st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
        guardado = st.form_submit_button("📦 Guardar inventario del día", type="primary",
                                         use_container_width=True)

    if guardado:
        comp_stock = {cid: (None if unl else qty) for cid, (unl, qty) in comp_vals.items()}
        menu_stock = {mid: (None if unl else qty) for mid, (unl, qty) in menu_vals.items()}
        guardar_inventario(comp_stock, menu_stock)
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Pestaña 7 · Importar menú desde Excel / CSV (con fusión al inventario)
# ══════════════════════════════════════════════════════════════════════════════
# Sube un .xlsx/.csv con una fila por ítem. La columna 'seccion' enruta cada fila a su
# panel (componentes del Plato del Día vs catálogo a la carta/especiales/bebidas) y se
# mapean nombre/descr/precio/stock de una sola vez. El upsert es NO destructivo: una
# celda de stock en blanco NUNCA toca el contador (un re-import de solo precios conserva
# el inventario), no duplica filas (identidad por nombre dentro de su sección) y no rompe
# el seguimiento en curso. Acceso: admin + caja (capacidad edit_menu).

# seccion (normalizada) → grupo del Plato del Día. Fijo: el importador solo reconoce
# los 4 grupos clásicos (los grupos dinámicos de plato_dia_grupos son otra pestaña y
# quedan fuera del alcance de este importador).
_SECCION_COMP = {
    "entrada": "entrada", "entradas": "entrada",
    "principio": "principio", "principios": "principio",
    "proteina": "proteina", "proteinas": "proteina",
    "carne": "proteina", "carnes": "proteina",
    "acompanamiento": "acompanamiento", "acompanamientos": "acompanamiento",
    "acomp": "acompanamiento", "guarnicion": "acompanamiento",
}
# Alias EXTRA (además de su propia etiqueta/clave) de las 4 categorías clásicas del
# catálogo, para no romper plantillas que los restaurantes ya vienen usando ('carta',
# 'fuerte', 'jugo'…). Cualquier categoría —clásica o creada por el restaurante en
# ⚙️ Ajustes → 🏷️ Categorías— SIEMPRE se reconoce además por su propia etiqueta y clave
# normalizadas (ver _seccion_map): un restaurante que crea 'Desayunos' puede escribir
# "desayunos" en la columna 'seccion' sin configurar nada aparte.
_ALIAS_MENU_CLASICOS = {
    "especial": "especial", "especiales": "especial",
    "a la carta": "a_la_carta", "a_la_carta": "a_la_carta",
    "carta": "a_la_carta", "plato": "a_la_carta",
    "plato a la carta": "a_la_carta", "fuerte": "a_la_carta",
    "adicional": "adicional", "adicionales": "adicional",
    "extra": "adicional", "extras": "adicional",
    "bebida": "bebida", "bebidas": "bebida",
    "jugo": "bebida", "gaseosa": "bebida", "drink": "bebida",
}


def _seccion_map() -> dict:
    """seccion normalizada → (clase, destino) para el importador. 'comp' =
    menu_componentes.grupo (los 4 clásicos); 'menu' = menu.categoria (cualquier
    categoría viva, clásica o nueva)."""
    m = {sec: ("comp", destino) for sec, destino in _SECCION_COMP.items()}
    m.update({sec: ("menu", destino) for sec, destino in _ALIAS_MENU_CLASICOS.items()})
    for c in cargar_categorias(solo_activas=False):
        m[_norm(c["clave"])] = ("menu", c["clave"])
        m[_norm(c["etiqueta"])] = ("menu", c["clave"])
    return m


# Alias aceptados por columna (normalizados: sin acentos, en minúscula).
COL_ALIAS = {
    "seccion":     ("seccion", "categoria", "tipo", "grupo", "seccion/categoria"),
    "nombre":      ("nombre", "name", "plato", "producto", "item", "articulo"),
    "descripcion": ("descripcion", "desc", "detalle"),
    "precio":      ("precio", "price", "valor", "costo"),
    "stock":       ("stock", "cantidad", "existencias", "inventario", "disponibles",
                    "cantidad disponible"),
    "activo":      ("activo", "active", "disponible", "habilitado"),
}
# Etiquetas fijas de los 4 grupos del Plato del Día que entiende el importador; el
# resto de destinos ('menu') se rotulan con etiquetas_categorias() (ver _dest_label).
_DEST_LABEL_COMP = {
    "entrada": "Entrada", "principio": "Principio", "proteina": "Proteína",
    "acompanamiento": "Acompañamientos",
}


def _dest_label(destino: str) -> str:
    """Etiqueta legible de un destino del importador: uno de los 4 grupos del Plato
    del Día, o la etiqueta actual de una categoría del catálogo (clásica o nueva)."""
    if destino in _DEST_LABEL_COMP:
        return _DEST_LABEL_COMP[destino]
    return etiquetas_categorias().get(destino, destino)


def _norm(s) -> str:
    """minúsculas + sin acentos + sin espacios extremos, para casar encabezados/secciones."""
    s = str("" if s is None else s).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def _parse_int(v):
    """int en [0, _INT32_MAX] desde celda numérica o texto ('25.000'/'$25,000' → 25000);
    None si está en blanco. Distingue número (uso directo) de texto (quita separadores de
    miles). Se acota al rango de INTEGER de Postgres para no abortar el import entero."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, bool):
        return None
    if isinstance(v, numbers.Number):
        try:
            return max(0, min(_INT32_MAX, int(round(float(v)))))
        except (TypeError, ValueError):
            return None
    s = str(v).strip().replace("$", "").replace(" ", "")
    if s == "":
        return None
    s = s.replace(".", "").replace(",", "")
    try:
        return max(0, min(_INT32_MAX, int(s)))
    except ValueError:
        return None


def _parse_activo(v) -> bool:
    """Activo por defecto; solo 'no/false/0/inactivo/n/off' lo desactivan."""
    try:
        if v is None or pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    return _norm(v) not in ("no", "false", "0", "inactivo", "n", "off")


def _celda_texto(v):
    """str limpia o None (para descripción)."""
    try:
        if v is None or pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    return s or None


def _mapear_columnas(df) -> dict:
    """{rol: nombre_real_de_columna} resolviendo alias por encabezado normalizado."""
    norm_cols = {_norm(c): c for c in df.columns}
    out = {}
    for rol, alias in COL_ALIAS.items():
        for a in alias:
            if a in norm_cols:
                out[rol] = norm_cols[a]
                break
    return out


def _parse_filas(df):
    """(filas_validas, errores). Cada fila: {kind, destino, nombre, descripcion, precio,
    stock, activo, fila}. Salta filas sin nombre; reporta secciones desconocidas."""
    cols = _mapear_columnas(df)
    if "seccion" not in cols or "nombre" not in cols:
        return [], ["Faltan columnas obligatorias: 'seccion' y 'nombre'. "
                    "Descarga la plantilla para ver el formato."]
    seccion_map = _seccion_map()
    filas, errores = [], []
    for i, row in df.iterrows():
        nfila = int(i) + 2  # +2: fila de encabezado + base 1
        nombre = _celda_texto(row[cols["nombre"]])
        if not nombre:
            continue  # fila vacía → se ignora en silencio
        sec_raw = row[cols["seccion"]]
        sec = _norm(sec_raw)
        if sec not in seccion_map:
            errores.append(f"Fila {nfila}: sección desconocida «{_celda_texto(sec_raw) or ''}» "
                           f"(ítem «{nombre}»).")
            continue
        kind, destino = seccion_map[sec]
        filas.append({
            "kind": kind, "destino": destino, "nombre": nombre,
            "descripcion": _celda_texto(row[cols["descripcion"]]) if "descripcion" in cols else None,
            "precio": _parse_int(row[cols["precio"]]) if "precio" in cols else None,
            "stock":  _parse_int(row[cols["stock"]]) if "stock" in cols else None,
            "activo": _parse_activo(row[cols["activo"]]) if "activo" in cols else True,
            "fila": nfila,
        })
    return filas, errores


def _leer_dataframe(archivo):
    """Lee el archivo subido a DataFrame. .csv nativo (con fallback latin-1); .xlsx vía
    openpyxl (ImportError si falta la librería → lo maneja la vista)."""
    nombre = (getattr(archivo, "name", "") or "").lower()
    if nombre.endswith(".csv"):
        try:
            return pd.read_csv(archivo, dtype=str)
        except UnicodeDecodeError:
            archivo.seek(0)
            return pd.read_csv(archivo, dtype=str, encoding="latin-1")
    # dtype=str también en .xlsx (paridad con el CSV): evita que un nombre numérico
    # ('100') llegue como float ('100.0') y que pandas infiera tipos raros; los parsers
    # (_parse_int/_parse_activo/_celda_texto) ya normalizan texto y blancos (NaN).
    return pd.read_excel(archivo, dtype=str)


def _calc_stock(actual, valor, sumar: bool, existe: bool):
    """Stock resultante de un upsert:
      · valor None (celda en blanco) → NO tocar (deja el actual; NULL/ilimitado si es nuevo).
      · sumar=False (sobrescribir)   → valor (inicializa el contador).
      · sumar=True  (reabastecer)    → actual + valor para los existentes con contador; el
                                       valor tal cual para nuevos o los que estaban ilimitados.
    Nunca 'borra' un contador: una celda en blanco preserva el estado vivo del inventario."""
    if valor is None:
        return actual if existe else None
    if sumar and existe and actual is not None:
        return int(actual) + int(valor)
    return int(valor)


def importar_menu(filas, modo_stock: str = "sobrescribir") -> dict:
    """Aplica el import en UNA transacción (atómico). Upsert por identidad:
    componentes por (grupo, LOWER(nombre)); catálogo por (categoria, LOWER(nombre)).
    No duplica filas ni rompe el inventario; el stock se fusiona según _calc_stock.
    Devuelve un resumen con conteos y omitidos."""
    res = {"comp_nuevos": 0, "comp_act": 0, "menu_nuevos": 0, "menu_act": 0,
           "omitidos": []}
    sumar = (modo_stock == "sumar")
    with engine.begin() as conn:
        for f in filas:
            nombre = f["nombre"]
            if f["kind"] == "comp":
                grupo = f["destino"]
                ex = conn.execute(text(
                    "SELECT id, stock FROM menu_componentes "
                    "WHERE grupo = :g AND LOWER(nombre) = LOWER(:n)"
                ), {"g": grupo, "n": nombre}).mappings().first()
                nuevo_stock = _calc_stock(ex["stock"] if ex else None, f["stock"], sumar, bool(ex))
                if ex:
                    conn.execute(text(
                        "UPDATE menu_componentes SET activo = :a, stock = :s WHERE id = :id"
                    ), {"a": f["activo"], "s": nuevo_stock, "id": ex["id"]})
                    res["comp_act"] += 1
                else:
                    mx = conn.execute(text(
                        "SELECT COALESCE(MAX(orden), 0) FROM menu_componentes WHERE grupo = :g"
                    ), {"g": grupo}).scalar()
                    conn.execute(text(
                        "INSERT INTO menu_componentes (grupo, nombre, activo, orden, stock) "
                        "VALUES (:g, :n, :a, :o, :s)"
                    ), {"g": grupo, "n": nombre, "a": f["activo"], "o": int(mx) + 1, "s": nuevo_stock})
                    res["comp_nuevos"] += 1
            else:
                categoria = f["destino"]
                ex = conn.execute(text(
                    "SELECT id, stock FROM menu WHERE categoria = :c AND LOWER(nombre) = LOWER(:n)"
                ), {"c": categoria, "n": nombre}).mappings().first()
                # Todas las categorías de catálogo (especial/a la carta/adicional/bebida)
                # usan el precio del archivo.
                precio = f["precio"]
                if not ex and (precio is None or precio <= 0):
                    res["omitidos"].append(
                        f"Fila {f['fila']}: «{nombre}» sin precio válido (requerido para "
                        f"un plato/bebida nuevo).")
                    continue
                nuevo_stock = _calc_stock(ex["stock"] if ex else None, f["stock"], sumar, bool(ex))
                if ex:
                    sets = ["activo = :a", "stock = :s"]
                    params = {"a": f["activo"], "s": nuevo_stock, "id": ex["id"]}
                    if precio is not None:             # precio en blanco → conserva el actual
                        sets.append("precio = :p"); params["p"] = int(precio)
                    if f["descripcion"] is not None:   # descr en blanco → conserva la actual
                        sets.append("descripcion = :d"); params["d"] = f["descripcion"]
                    conn.execute(text(f"UPDATE menu SET {', '.join(sets)} WHERE id = :id"), params)
                    res["menu_act"] += 1
                else:
                    mx = conn.execute(text(
                        "SELECT COALESCE(MAX(orden), 0) FROM menu WHERE categoria = :c"
                    ), {"c": categoria}).scalar()
                    conn.execute(text(
                        "INSERT INTO menu (nombre, precio, activo, orden, categoria, descripcion, stock) "
                        "VALUES (:n, :p, :a, :o, :c, :d, :s)"
                    ), {"n": nombre, "p": int(precio or 0), "a": f["activo"], "o": int(mx) + 1,
                        "c": categoria, "d": f["descripcion"], "s": nuevo_stock})
                    res["menu_nuevos"] += 1
    _clear_menu_caches()
    cargar_componentes.clear()
    return res


def _plantilla_csv() -> bytes:
    """Plantilla de ejemplo descargable (utf-8-sig para que Excel abra los acentos)."""
    filas = [
        ["seccion", "nombre", "descripcion", "precio", "stock", "activo"],
        ["entrada", "Sopa del día", "", "", "50", "si"],
        ["principio", "Frijol", "", "", "40", "si"],
        ["proteina", "Pollo", "", "", "40", "si"],
        ["acompanamiento", "Arroz", "", "", "100", "si"],
        ["especial", "Bandeja Paisa", "Frijol, arroz, carne, chicharrón, huevo", "22000", "20", "si"],
        ["a_la_carta", "Hamburguesa clásica", "", "25000", "30", "si"],
        ["bebida", "Limonada natural", "", "5000", "", "si"],
    ]
    buf = io.StringIO()
    csv.writer(buf).writerows(filas)
    return buf.getvalue().encode("utf-8-sig")


def _plantilla_secciones() -> list:
    """Opciones del desplegable de 'seccion' en la plantilla Excel (texto amigable; el
    parser las normaliza vía _seccion_map): los 4 grupos clásicos del Plato del Día +
    la etiqueta de cada categoría ACTIVA del catálogo (clásicas y las que el
    restaurante haya agregado en ⚙️ Ajustes → 🏷️ Categorías). El formula1 de Excel
    tiene un límite práctico de ~255 caracteres: con muchas categorías de nombre largo
    el desplegable puede quedar incompleto (el importador igual las reconoce si se
    escriben a mano)."""
    return (["Entrada", "Principio", "Proteína", "Acompañamiento"]
            + [c["etiqueta"] for c in cargar_categorias()])


# Ejemplos (uno por sección) que se precargan en la hoja, en estilo tenue para que el
# cliente los reemplace. [Sección, Nombre, Descripción, Precio, Stock, Activo].
PLANTILLA_EJEMPLOS = [
    ["Entrada", "Sopa del día", None, None, 50, "Sí"],
    ["Principio", "Frijol", None, None, 40, "Sí"],
    ["Proteína", "Pollo a la plancha", None, None, 40, "Sí"],
    ["Acompañamiento", "Arroz blanco", None, None, 100, "Sí"],
    ["Especiales", "Bandeja Paisa", "Frijol, arroz, carne, chicharrón, huevo, plátano",
     22000, 20, "Sí"],
    ["A la carta", "Hamburguesa clásica", "Carne 150g, queso y papas", 25000, 30, "Sí"],
    ["Bebidas", "Limonada natural", "Vaso 16oz", 5000, None, "Sí"],
]


def _plantilla_xlsx() -> bytes:
    """Plantilla Excel rica: encabezados con estilo, desplegables (Sección / Activo),
    comentarios de ayuda, formato de precio/stock, ejemplos por sección y una hoja de
    Instrucciones. Mantiene las columnas que el importador entiende. Requiere openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    AZUL, AZUL_CLARO, GRIS, GRIS_TXT = "1E293B", "F1F5F9", "F8FAFC", "64748B"
    wb = Workbook()

    # ── Hoja 1 · Menú (la que llena el cliente) ─────────────────────────────────
    ws = wb.active
    ws.title = "Menú"
    encabezados = ["Sección", "Nombre", "Descripción", "Precio", "Stock", "Activo"]
    ws.append(encabezados)

    fill_h = PatternFill("solid", fgColor=AZUL)
    font_h = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    borde = Border(bottom=Side(style="thin", color="CBD5E1"))
    for col in range(1, len(encabezados) + 1):
        c = ws.cell(row=1, column=col)
        c.fill, c.font, c.alignment, c.border = fill_h, font_h, center, borde
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    for i, w in enumerate([16, 30, 44, 12, 10, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ayudas = {
        1: ("Sección del menú (elige de la lista). Entrada · Principio · Proteína · "
            "Acompañamiento van al PLATO DEL DÍA. Especiales · A la carta · Bebidas van a la carta."),
        2: "Nombre del ítem. Si ya existe en esa sección, se ACTUALIZA (no se duplica).",
        3: "Opcional. Lo que incluye (sobre todo útil en Especiales).",
        4: ("Precio en pesos, solo el número (ej: 25000). Especiales, A la carta y Bebidas: "
            "obligatorio si es nuevo. Componentes del Plato del Día: sin precio."),
        5: "Cantidad de hoy. VACÍO = ilimitado / no cambia el contador. 0 = agotado.",
        6: "Sí = se ofrece · No = inactivo.",
    }
    for col, txt in ayudas.items():
        cm = Comment(txt, "Plantilla")
        cm.width, cm.height = 280, 130
        ws.cell(row=1, column=col).comment = cm

    # Ejemplos en estilo tenue (el cliente los reemplaza).
    font_ej = Font(italic=True, color=GRIS_TXT)
    fill_ej = PatternFill("solid", fgColor=GRIS)
    for fila in PLANTILLA_EJEMPLOS:
        ws.append(fila)
    n_ej = len(PLANTILLA_EJEMPLOS)
    for row in range(2, 2 + n_ej):
        for col in range(1, 7):
            ws.cell(row=row, column=col).font = font_ej
            ws.cell(row=row, column=col).fill = fill_ej
    nota = Comment("Filas de ejemplo: edítalas o bórralas y escribe tu propio menú.", "Plantilla")
    nota.width, nota.height = 260, 80
    ws.cell(row=2, column=2).comment = nota

    # Formato numérico + relleno alterno para muchas filas vacías listas para llenar.
    ultima = 400
    for row in range(2, ultima + 1):
        ws.cell(row=row, column=4).number_format = "#,##0"
        ws.cell(row=row, column=5).number_format = "0"

    # Desplegables (validación de datos) en Sección y Activo.
    dv_sec = DataValidation(type="list",
                            formula1='"' + ",".join(_plantilla_secciones()) + '"',
                            allow_blank=True)
    dv_sec.errorTitle, dv_sec.error = "Sección inválida", "Elige una sección de la lista."
    dv_sec.promptTitle, dv_sec.prompt = "Sección", "Elige a qué panel va este ítem."
    ws.add_data_validation(dv_sec)
    dv_sec.add(f"A2:A{ultima}")

    dv_act = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    ws.add_data_validation(dv_act)
    dv_act.add(f"F2:F{ultima}")

    # ── Hoja 2 · Instrucciones ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instrucciones")
    ws2.column_dimensions["A"].width = 115
    wrap = Alignment(wrap_text=True, vertical="top")
    f_title = Font(bold=True, size=16, color=AZUL)
    f_h = Font(bold=True, size=12, color=AZUL)
    f_p = Font(size=11, color="334155")
    lineas = [
        ("t", "📥 Cómo llenar la plantilla del menú"),
        ("p", "Llena la hoja «Menú»: una fila por cada plato, ingrediente o bebida. "
              "Al subirla verás una vista previa antes de guardar."),
        ("h", "Sección (obligatoria · lista desplegable)"),
        ("p", "• Entrada / Principio / Proteína / Acompañamiento → componentes del Plato del Día."),
        ("p", "• Especiales → platos especiales con su propio precio (y descripción de lo que incluyen)."),
        ("p", "• A la carta → platos sueltos con su propio precio."),
        ("p", "• Bebidas → bebidas con su propio precio."),
        ("h", "Nombre (obligatorio)"),
        ("p", "Si el nombre ya existe en esa sección, se ACTUALIZA; no se crea un duplicado."),
        ("h", "Descripción (opcional)"),
        ("p", "Lo que incluye el plato. Recomendado en Especiales."),
        ("h", "Precio"),
        ("p", "Solo el número, en pesos (ej: 25000). Obligatorio para Especiales, A la carta y "
              "Bebidas nuevas. Componentes del Plato del Día no llevan precio."),
        ("h", "Stock (inventario del día)"),
        ("p", "• Un número = porciones/unidades disponibles hoy."),
        ("p", "• Vacío = ilimitado y NO cambia el contador actual (ideal para actualizar solo precios)."),
        ("p", "• 0 = agotado."),
        ("h", "Activo"),
        ("p", "Sí = se ofrece · No = queda inactivo."),
        ("h", "Al subir"),
        ("p", "• «Sobrescribir» fija el stock al número del archivo · «Sumar» lo añade al actual."),
        ("p", "• Una celda de stock vacía nunca borra el inventario en curso."),
        ("p", "• Borra o reemplaza las filas de ejemplo de la hoja «Menú»."),
    ]
    for r, (kind, txt) in enumerate(lineas, start=1):
        c = ws2.cell(row=r, column=1, value=txt)
        c.alignment = wrap
        c.font = f_title if kind == "t" else (f_h if kind == "h" else f_p)
        if kind == "t":
            ws2.row_dimensions[r].height = 26

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_importar():
    # Resultado del import previo (sobrevive al reset del uploader): se muestra y se limpia.
    prev = st.session_state.pop("import_resultado", None)
    if prev:
        nuevos = prev["comp_nuevos"] + prev["menu_nuevos"]
        act = prev["comp_act"] + prev["menu_act"]
        st.success(f"✅ Importación completada · {nuevos} ítem(s) nuevo(s) · {act} actualizado(s).")
        if prev["omitidos"]:
            with st.expander(f"⚠️ {len(prev['omitidos'])} fila(s) omitida(s)"):
                for o in prev["omitidos"]:
                    st.markdown(f"- {html.escape(o)}")

    # Las secciones válidas se listan desde las categorías VIVAS (no una lista quemada, que
    # se quedaba corta: omitía Adicionales y toda categoría creada por el restaurante).
    secciones_txt = " · ".join(
        ["entrada", "principio", "proteina", "acompanamiento"]
        + [c["etiqueta"] for c in cargar_categorias(solo_activas=False)])
    st.markdown(
        '<div style="background:#fafaf8; border:1px solid #ececec; border-radius:10px; '
        'padding:0.8rem 1rem; font-size:0.85rem; color:#45443e; margin-bottom:1rem;">'
        '📥 Sube el menú en <b>.xlsx</b> o <b>.csv</b>, una fila por ítem. La columna '
        '<b>seccion</b> enruta cada fila a su panel y se cargan nombre, descripción, precio '
        'y stock de una vez. Si una fila ya existe (mismo nombre en su sección) se '
        '<b>actualiza</b> sin duplicar; una celda de <b>stock en blanco no cambia</b> el '
        'inventario en curso.<br><span style="color:#6b6b64;">Secciones válidas: '
        f'{html.escape(secciones_txt)}.</span></div>',
        unsafe_allow_html=True,
    )

    # Plantilla rica en Excel (desplegables + ejemplos + hoja de instrucciones); .csv simple
    # como respaldo (y por si el servidor aún no tiene openpyxl).
    dcol1, dcol2 = st.columns(2)
    with dcol1:
        try:
            st.download_button(
                "⬇️ Descargar plantilla Excel", data=_plantilla_xlsx(),
                file_name="plantilla_menu.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary",
                help="Excel guiado: listas desplegables, ejemplos y hoja de instrucciones.")
        except Exception:
            st.caption("Plantilla Excel no disponible en este servidor; usa la versión .csv →")
    with dcol2:
        st.download_button(
            "⬇️ Plantilla simple (.csv)", data=_plantilla_csv(),
            file_name="plantilla_menu.csv", mime="text/csv", use_container_width=True,
            help="Misma estructura en texto plano, por si prefieres CSV.")

    nonce = int(st.session_state.get("import_nonce", 0))
    archivo = st.file_uploader("Archivo del menú (.xlsx o .csv)", type=["xlsx", "csv"],
                               key=f"import_file_{nonce}")
    if archivo is None:
        return

    try:
        df = _leer_dataframe(archivo)
    except ImportError:
        st.error("No se pudo leer el .xlsx (falta la librería openpyxl en el servidor). "
                 "Mientras tanto, sube el menú en formato .csv.")
        return
    except Exception as e:  # archivo corrupto / formato inesperado
        st.error(f"No se pudo leer el archivo: {html.escape(str(e))}")
        return

    if df is None or df.empty:
        st.warning("El archivo no tiene filas.")
        return

    filas, errores = _parse_filas(df)

    st.markdown('<div class="section-title">Vista previa</div>', unsafe_allow_html=True)
    # st.dataframe serializa a Arrow y puede lanzar con columnas de tipos mixtos/raros.
    # La vista previa es cosmética: si falla, no debe bloquear la importación.
    try:
        st.dataframe(df, use_container_width=True, height=240)
    except Exception:
        st.caption("Vista previa no disponible para este archivo; la importación sí puede continuar.")

    if errores:
        st.warning("Avisos:\n" + "\n".join(f"- {e}" for e in errores))

    if not filas:
        st.error("No hay filas válidas para importar. Revisa la columna 'seccion' y 'nombre'.")
        return

    # Conteo por destino (qué panel recibe qué).
    conteo = {}
    for f in filas:
        conteo[f["destino"]] = conteo.get(f["destino"], 0) + 1
    chips = " ".join(
        f'<span class="badge badge-activo">{_dest_label(d)}: {n}</span>'
        for d, n in sorted(conteo.items()))
    st.markdown(f'<div style="margin:6px 0 10px;">{chips}</div>', unsafe_allow_html=True)

    st.caption("Stock al importar")
    modo = st.radio("Stock al importar", ["Sobrescribir (inicializar)", "Sumar (reabastecer)"],
                    horizontal=True, label_visibility="collapsed", key="import_modo_stock")
    modo_stock = "sumar" if modo.startswith("Sumar") else "sobrescribir"
    st.caption("Sobrescribir fija el stock al número del archivo; Sumar lo añade al actual. "
               "En ambos casos, una celda vacía deja el contador como está.")

    if st.button(f"📥 Importar {len(filas)} ítem(s)", type="primary",
                 use_container_width=True, key="btn_importar_menu"):
        # El import es atómico (engine.begin): si algo revienta, NO queda nada a medias.
        # Capturamos el fallo para mostrar un mensaje claro en vez de un stacktrace rojo.
        try:
            res = importar_menu(filas, modo_stock)
        except Exception as e:
            st.error(f"No se pudo importar (no se aplicó ningún cambio): {html.escape(str(e))}")
            return
        st.session_state["import_resultado"] = res
        st.session_state["import_nonce"] = nonce + 1   # limpia el uploader y la vista previa
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN: MENÚ
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# VISTA DE SOLO LECTURA (mesero) — carta activa de hoy, sin edición
# ══════════════════════════════════════════════════════════════════════════════
def _render_readonly():
    """Carta de SOLO consulta para el mesero: secciones, platos disponibles HOY y
    precios. Sin pestañas de edición ni modales de crear/editar/eliminar/86."""
    st.markdown(titulo_seccion('🍔 Carta (solo lectura)'),
                unsafe_allow_html=True)

    def _cards(rows):
        for r in rows:
            nombre = html.escape(_txt(r.get("nombre")))
            desc = html.escape(_txt(r.get("descripcion")))
            precio = int(r.get("precio", 0) or 0)
            sin_stock = agotado_por_stock(r.get("stock"))
            card_cls = "menu-card inactivo" if sin_stock else "menu-card"
            desc_html = f'<div class="menu-precio">{desc}</div>' if desc else ""
            badge = (' <span class="badge badge-agotado">🚫 Sin stock</span>'
                     if sin_stock else "")
            st.markdown(
                f'<div class="{card_cls}">'
                f'<div style="display:flex; justify-content:space-between; align-items:center;">'
                f'<div><div class="menu-nombre">{nombre}</div>{desc_html}'
                f'<div class="menu-precio">${fmt_money(precio)}</div></div>'
                f'<div style="text-align:right;">{_stock_chip(r.get("stock"))}{badge}</div>'
                f'</div></div>',
                unsafe_allow_html=True,
            )

    # 🍽️ Plato del Día (precio plano + componentes disponibles por grupo dinámico)
    grupos_ro = cargar_grupos_pd()
    reglas_ro = " · ".join(f"{g['etiqueta'].lower()}: {_regla_grupo(g)}"
                           for g in grupos_ro if g["max_sel"] > 1 or g["min_sel"] == 0)
    st.markdown(
        f'<div style="background:#fafaf8; border:1px solid #ececec; border-radius:10px; '
        f'padding:0.7rem 1rem; font-size:0.85rem; color:#45443e; margin:0.5rem 0 1rem 0;">'
        f'🍽️ <b>Plato del Día</b> · ${fmt_money(precio_plato_dia())}'
        f'{" · " + html.escape(reglas_ro) if reglas_ro else ""}</div>',
        unsafe_allow_html=True,
    )
    def _opt_label(o):
        """Nombre del componente con sus porciones para el mesero: 'Pollo (12)',
        ámbar si quedan pocas, tachado si está agotado, sin nada si es ilimitado."""
        s = stock_int(o.get("stock"))
        nom = html.escape(str(o["nombre"]))
        if s is None:
            return nom
        if s <= 0:
            return (f'<span style="color:#b91c1c; text-decoration:line-through;">'
                    f'{nom} (agotado)</span>')
        color = "#b45309" if s <= STOCK_BAJO else "#6b6b64"
        return f'{nom} <span style="color:{color}; font-weight:600;">({s})</span>'

    por_grupo = componentes_activos_por_grupo()
    etiquetas_ro = etiquetas_grupos_pd()
    for grupo in por_grupo:
        opciones = por_grupo.get(grupo, [])
        if not opciones:
            continue
        nombres = " · ".join(_opt_label(o) for o in opciones)
        st.markdown(
            f'<div style="margin-bottom:0.6rem;"><span style="font-weight:600; color:#26262b;">'
            f'{html.escape(etiquetas_ro.get(grupo, grupo))}:</span> '
            f'<span style="color:#6b6b64; font-size:0.88rem;">{nombres}</span></div>',
            unsafe_allow_html=True,
        )

    cat = cargar_catalogo()
    disp = disponibles(cat) if not cat.empty else cat

    def _seccion(titulo: str, categoria: str):
        rows = (disp[disp["categoria"] == categoria].to_dict("records")
                if not disp.empty else [])
        st.markdown(titulo_seccion(titulo), unsafe_allow_html=True)
        if rows:
            _cards(rows)
        else:
            st.markdown('<p style="color:#a3a39b; font-size:0.85rem;">Sin platos activos.</p>',
                        unsafe_allow_html=True)

    for c in cargar_categorias():
        _seccion(f'{c["emoji"]} {c["etiqueta"]}'.strip(), c["clave"])


def _filtrar_seccion(sub, key: str, placeholder: str):
    """Buscador POR SECCIÓN: dibuja un text_input de filtro y devuelve el sub-DataFrame
    con solo las filas cuyo nombre (o descripción, si existe) casa el término. Insensible
    a acentos/mayúsculas (reusa _norm). Sin término devuelve 'sub' tal cual. Cada grupo del
    Plato del Día y cada categoría de catálogo tiene el suyo, para no recorrer la lista
    entera cuando el menú es grande. Devuelve (sub_filtrado, hubo_termino)."""
    if sub is None or sub.empty:
        return sub, False
    term = st.text_input("Buscar", key=key, placeholder=placeholder,
                         label_visibility="collapsed")
    q = _norm((term or "").strip())
    if not q:
        return sub, False
    mask = sub["nombre"].apply(lambda n: q in _norm(_txt(n)))
    if "descripcion" in sub.columns:
        mask = mask | sub["descripcion"].apply(lambda d: q in _norm(_txt(d)))
    return sub[mask], True


def render():
    # RBAC: el mesero solo CONSULTA la carta. Bloquea todas las pestañas de edición,
    # los modales (crear/editar/eliminar) y los toggles (activo / 86).
    if not auth.can("edit_menu"):
        _render_readonly()
        return

    # Las CATEGORÍAS siguen siendo las pestañas de arriba, ahora leídas de la tabla
    # categorias (cada restaurante agrega las suyas en ⚙️ Ajustes → 🏷️ Categorías) en vez
    # de una lista fija. Lo que se pliega/despliega son los PLATOS dentro de cada
    # categoría: cada catálogo los agrupa en dos acordeones Disponibles / No disponibles
    # (ver _render_catalogo_tab). Inventario e Importar se movieron a 💰 Caja; aquí el
    # Menú se enfoca en la carta y los ajustes.
    _inject_accordion_css()
    categorias = cargar_categorias()
    labels = (["🍽️ Plato del Día"]
              + [f'{c["emoji"]} {c["etiqueta"]}'.strip() for c in categorias]
              + ["⚙️ Ajustes"])
    tabs = st.tabs(labels)
    with tabs[0]:
        _render_plato_dia()
    for tab, c in zip(tabs[1:-1], categorias):
        with tab:
            comport = comportamiento_categoria(c["clave"])
            _render_catalogo_tab(c["clave"], c["etiqueta"], con_precio=True,
                                 con_desc=comport["desc"])
    with tabs[-1]:
        _render_ajustes()
